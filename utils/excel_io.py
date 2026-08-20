"""Excel 读写工具"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def read_excel(filepath):
    """读取Excel，返回 (DataFrame, 列名列表)"""
    df = pd.read_excel(filepath, engine="openpyxl", dtype=str)
    df = df.fillna("")
    return df, list(df.columns)


def safe_get(row, col_name, default=""):
    """安全获取单元格值"""
    val = row.get(col_name, default)
    if pd.isna(val):
        return default
    return str(val)


RISK_FILLS = {
    "高风险有其他违规风险": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
    "中风险只警告不扣分": PatternFill(start_color="FFF5CC", end_color="FFF5CC", fill_type="solid"),
    "低风险无违规风险": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
}
RISK_FONTS = {
    "高风险有其他违规风险": Font(bold=True, color="CC0000"),
    "中风险只警告不扣分": Font(bold=True, color="996600"),
    "低风险无违规风险": Font(color="006600"),
}


def write_excel(df, output_path):
    """输出带颜色标记和自动换行的Excel"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "RiskAnalysis"

    # 写表头
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 找风险相关列索引
    risk_col_name = "风险评级"
    text_col_names = ["高风险原因", "高风险补救措施", "中风险原因", "中风险补救措施"]

    def _col_idx(name):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    risk_col_idx = _col_idx(risk_col_name)
    text_col_indices = [i for i in (_col_idx(n) for n in text_col_names) if i]

    # 写数据行
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(headers, 1):
            val = str(row.get(col_name, "")) if not pd.isna(row.get(col_name, "")) else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

        # 风险列颜色标记
        if risk_col_idx:
            risk_cell = ws.cell(row=row_idx, column=risk_col_idx)
            risk_val = str(risk_cell.value)
            fill = RISK_FILLS.get(risk_val)
            fnt = RISK_FONTS.get(risk_val)
            if fill:
                risk_cell.fill = fill
            if fnt:
                risk_cell.font = fnt

        # 原因和补救措施列自动换行
        for col_idx in text_col_indices:
            c = ws.cell(row=row_idx, column=col_idx)
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # 设置列宽
    col_widths = {}
    if risk_col_idx:
        col_widths[risk_col_idx] = 22
    for col_idx in text_col_indices:
        col_widths[col_idx] = 60
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 其它列自动宽度（限制最大50）
    for col_idx in range(1, len(headers) + 1):
        if col_idx not in col_widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(50, len(str(headers[col_idx - 1])) + 4))

    wb.save(output_path)
    return output_path
