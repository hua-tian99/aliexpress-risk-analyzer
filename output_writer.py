"""
输出写入器 — 将分析结果写回Excel并格式化
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from utils.excel_io import write_excel as write_excel_base
from risk_evaluator import RiskEvaluator


class OutputWriter:
    """输出写入器 — 聚合结果并写入Excel"""

    RISK_FILLS = {
        "高风险有其他违规风险": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
        "中风险只警告不扣分": PatternFill(start_color="FFF5CC", end_color="FFF5CC", fill_type="solid"),
        "低风险无违规风险": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
    }
    RISK_FONTS = {
        "高风险有其他违规风险": Font(bold=True, color="CC0000"),
        "中风险只警告不扣分": Font(bold=True, color="996600"),
        "低风险无违规风险": Font(color="006600"),
    }

    def __init__(self, df, all_results):
        """
        Args:
            df: pd.DataFrame, 原始Excel数据
            all_results: dict[行索引 -> list[detector_result]], 所有产品的检测结果
        """
        self.df = df.copy()
        self.all_results = all_results

    def add_columns(self):
        """添加风险评级 + 高风险原因/补救措施 + 中风险原因/补救措施 共五列"""
        risk_ratings = []
        high_reasons = []
        high_remedies = []
        medium_reasons = []
        medium_remedies = []

        for idx, row in self.df.iterrows():
            results = self.all_results.get(idx, [])
            evaluator = RiskEvaluator(results)
            risk_ratings.append(evaluator.get_overall_risk())
            high_reasons.append(evaluator.format_high_risk_reasons())
            high_remedies.append(evaluator.format_high_risk_remedies())
            medium_reasons.append(evaluator.format_medium_risk_reasons())
            medium_remedies.append(evaluator.format_medium_risk_remedies())

        self.df["风险评级"] = risk_ratings
        self.df["高风险原因"] = high_reasons
        self.df["高风险补救措施"] = high_remedies
        self.df["中风险原因"] = medium_reasons
        self.df["中风险补救措施"] = medium_remedies
        return self.df

    def save(self, output_path):
        """保存为格式化的Excel文件"""
        self.add_columns()
        return write_excel_base(self.df, output_path)

    def get_stats(self):
        """获取统计信息"""
        self.add_columns()
        high = (self.df["风险评级"] == "高风险有其他违规风险").sum()
        medium = (self.df["风险评级"] == "中风险只警告不扣分").sum()
        low = (self.df["风险评级"] == "低风险无违规风险").sum()

        # 统计各违规类别出现次数
        category_counts = {}
        for idx, results in self.all_results.items():
            for r in results:
                cat = r.get("category", "其他")
                if cat not in category_counts:
                    category_counts[cat] = 0
                category_counts[cat] += 1

        return {
            "total": len(self.df),
            "high": int(high),
            "medium": int(medium),
            "low": int(low),
            "category_counts": dict(sorted(category_counts.items(), key=lambda x: -x[1])),
        }
